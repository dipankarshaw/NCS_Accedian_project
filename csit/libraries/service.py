#!/usr/local/bin/python3

import time
import json
import os
import sys
import yaml
import re
from pprint import pprint
from netmiko import Netmiko,ConnectHandler
import datetime
from tqdm import tqdm
from jinja2 import Template
import csv
import textfsm
from multiprocessing import Pool
from ttp import ttp
import concurrent.futures as cf
from os import listdir
from ncclient import manager
import xmltodict
from collections import OrderedDict
from pygnmi.client import gNMIclient
import requests
import sys
import random
from openpyxl import load_workbook

create_delete_list = ['create','delete']
Loop_list = ['L1','L2']
mep_meg_dmm_slm_list = ['meg','mep','dmm','slm']
maxhosts = 5

file_path = os.path.dirname(os.path.realpath(__file__))
test_result = {}
dict3 = {}

class Jinja_user_func:
    def __init__(self):
        self
    def add_2_percent(self,service_BW,main_interface_bw):
        '''add 2 % extra bw in the Policy when 102 % of service BW is less than Main interface BW'''
        if service_BW * 102 //100 >= main_interface_bw:
            return main_interface_bw
        else:
            return service_BW * 102 // 100
    def add_13_percent(self,service_BW,main_interface_bw):
        '''add 2 % extra bw in the Policy when 102 % of service BW is less than Main interface BW'''
        if service_BW * 113 //100 >= main_interface_bw:
            return main_interface_bw
        else:
            return service_BW * 113 // 100
    def add_13_percent(self,service_BW,main_interface_bw):
        '''add 2 % extra bw in the Policy when 102 % of service BW is less than Main interface BW'''
        if service_BW * 113 //100 >= main_interface_bw:
            return main_interface_bw
        else:
            return service_BW * 113 // 100
    def lag_peak_percent_calc(self,service_BW,Lag_bw):
        ''' if percentage calculation of service BW w.r.t lag BW, comes in fraction then add move to next integer '''
        if service_BW * 100 % Lag_bw > 0:
            return service_BW * 100 // Lag_bw + 1
        else:
            return service_BW * 100 // Lag_bw
    def lag_CIR_percent_calc(self,service_BW,CIR_percentage,Lag_bw):
        ''' if CIR percentage calculation of service BW w.r.t lag BW, comes in fraction then add move to next integer '''
        if service_BW * CIR_percentage % Lag_bw > 0:
            return service_BW * CIR_percentage // Lag_bw + 1
        else:
            return service_BW * CIR_percentage// Lag_bw
    def peak_percent_calc(self,service_BW,main_interface_bw):
        ''' if percentage calculation of service BW w.r.t main interface BW, comes in fraction then add move to next integer '''
        if service_BW * 100 % main_interface_bw > 0:
            return service_BW * 100 // main_interface_bw + 1
        else:
            return service_BW * 100 // main_interface_bw
    def HQOS_Burst_calc(self,distributed_service_BW,main_interface_bw,QOS_type,percent,wb):
        ''' calculte the CBS value for differnt HQOS classes'''
        for row in wb[QOS_type].rows:
            if main_interface_bw == row[0].value and distributed_service_BW * percent // 100 == row[1].value:
                return row[2].value
            elif main_interface_bw == row[0].value and distributed_service_BW * percent // 100 > row[1].value and distributed_service_BW * percent // 100 < row[4].value:
                return row[2].value       
    def lag_13_per_extra_calc(self,service_BW,Lag_bw):
        ''' if 113 % of service BW is greater than LAG BW, then return what is service BW percentage of LAG BW
        else, if 113 % of service BW comes to be an Float then Move to next Integer
        else, return percentage of 113% of Service BW WRT LAG BW '''
        if service_BW * 113 // 100 >= Lag_bw:
            return service_BW * 100 // Lag_bw
        else:
            if service_BW * 113 % Lag_bw > 0:
                return service_BW * 113 // Lag_bw + 1
            else:
                return service_BW * 113 // Lag_bw     



class Service:

    def __init__(self,**kwargs):
        self.data = kwargs

    def connect_nodes(self):
        ''' Purpose:
        - Connect all cisco/accedian nodes via netmiko and stores an netmiko connection object in the Dictionary
        - Connect all Cisco nodes with Netconf & stores an netconf connection object in the dictionary
        - connects to H-Tool if SRTE is True
        '''
        for node in self.data["site_list"]:
            node['connect_obj'] = Netmiko(keepalive = 40, banner_timeout = 60, **node['login'])
            print("**** Netmiko connection established with {}".format(node['Node_name']))
        if self.data['SRTE']:
            self.data['Htool']['connect_obj'] = Netmiko(**self.data['Htool']['login'])
            print("**** Netmiko connection established with {}".format(self.data['Htool']['login']['host']))
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr':
                node['net_conf_obj'] = manager.connect(host = node['login']['host'], port = 830 , username = node['login']['username'], password = node['login']['password'],device_params={'name':'iosxr'}, hostkey_verify=False )
                print(f"**** Netconf connection established with {node['Node_name']}")
    def disconnect_nodes(self):
        ''' Purpose:
        - release netmiko Connections.
        - release netconf connections.
        - release the H-Tool connection if SRTE is True.
        '''        
        for node in self.data["site_list"]:
            node['connect_obj'].disconnect()
            print("**** Netmiko connection released from {}".format(node['Node_name']))
        if self.data['SRTE']:
            self.data['Htool']['connect_obj'].disconnect()
            print("**** Netmiko connection released from {}".format(self.data['Htool']['login']['host']))
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr':
                node['net_conf_obj'].close_session()
                print(f"**** Netconf connection released from {node['Node_name']}")           
    def connect_OLO_nodes(self):
        for olo in self.data["OLO_site_list"]:
            olo['connect_obj'] = Netmiko(**olo['login'])
            print("**** connection established with node {}".format(olo['Node_name']))
    def disconnect_OLO_nodes(self):
        for olo in self.data["OLO_site_list"]:
            olo['connect_obj'].disconnect()
            print("**** disconnected successfully from node {}".format(olo['Node_name']))
    def create_spirent_input_dict(self):
        '''
        purpose:
        - to create the dictionary to be fed to Spirent code
        '''

        dict10 = {}
        dict10['Spirent_2TAG_AZ'] = {}
        dict10['Spirent_2TAG_ZA'] = {}
        dict10['Spirent_1TAG_AZ'] = {}
        dict10['Spirent_1TAG_ZA'] = {}
        dict10['Spirent_0TAG_AZ'] = {}
        dict10['Spirent_0TAG_ZA'] = {}

        dict10['Spirent_2TAG_AZ']['UC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000,'Outer_VLAN_ID': str(self.data['item'] + 100), 'Inner_VLAN_ID': str(self.data['item']),'MAC_Dest':'00:10:94:02:00:01','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_2TAG_AZ']['BC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000,'Outer_VLAN_ID': str(self.data['item'] + 100), 'Inner_VLAN_ID': str(self.data['item']),'MAC_Dest':'FF:FF:FF:FF:FF:FF','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_2TAG_AZ']['MC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000,'Outer_VLAN_ID': str(self.data['item'] + 100), 'Inner_VLAN_ID': str(self.data['item']),'MAC_Dest':'01:00:5E:0B:01:02','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_2TAG_ZA']['UC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000,'Outer_VLAN_ID': str(self.data['item'] + 100), 'Inner_VLAN_ID': str(self.data['item']),'MAC_Dest':'00:10:94:01:00:01','MAC_Src': '00:10:94:02:00:01'}
        dict10['Spirent_2TAG_ZA']['BC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000,'Outer_VLAN_ID': str(self.data['item'] + 100), 'Inner_VLAN_ID': str(self.data['item']),'MAC_Dest':'FF:FF:FF:FF:FF:FF','MAC_Src': '00:10:94:02:00:01'}
        dict10['Spirent_2TAG_ZA']['MC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000,'Outer_VLAN_ID': str(self.data['item'] + 100), 'Inner_VLAN_ID': str(self.data['item']),'MAC_Dest':'01:00:5E:0B:01:02','MAC_Src': '00:10:94:02:00:01'}

        dict10['Spirent_1TAG_AZ']['UC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000,'VLAN_ID': str(self.data['item']), 'MAC_Dest':'00:10:94:02:00:01','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_1TAG_AZ']['BC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000,'VLAN_ID': str(self.data['item']), 'MAC_Dest':'FF:FF:FF:FF:FF:FF','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_1TAG_AZ']['MC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000,'VLAN_ID': str(self.data['item']), 'MAC_Dest':'01:00:5E:0B:01:02','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_1TAG_ZA']['UC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000,'VLAN_ID': str(self.data['item']), 'MAC_Dest':'00:10:94:01:00:01','MAC_Src': '00:10:94:02:00:01'}
        dict10['Spirent_1TAG_ZA']['BC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000,'VLAN_ID': str(self.data['item']), 'MAC_Dest':'FF:FF:FF:FF:FF:FF','MAC_Src': '00:10:94:02:00:01'}
        dict10['Spirent_1TAG_ZA']['MC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000,'VLAN_ID': str(self.data['item']), 'MAC_Dest':'01:00:5E:0B:01:02','MAC_Src': '00:10:94:02:00:01'}


        dict10['Spirent_0TAG_AZ']['UC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000, 'MAC_Dest':'00:10:94:02:00:01','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_0TAG_AZ']['BC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000, 'MAC_Dest':'FF:FF:FF:FF:FF:FF','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_0TAG_AZ']['MC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000, 'MAC_Dest':'01:00:5E:0B:01:02','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_0TAG_ZA']['UC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000, 'MAC_Dest':'00:10:94:01:00:01','MAC_Src': '00:10:94:02:00:01'}
        dict10['Spirent_0TAG_ZA']['BC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000, 'MAC_Dest':'FF:FF:FF:FF:FF:FF','MAC_Src': '00:10:94:02:00:01'}
        dict10['Spirent_0TAG_ZA']['MC'] = {'Rate_Mbps': (self.data['service_BW']*self.data['STP_percentage'])//100000, 'MAC_Dest':'01:00:5E:0B:01:02','MAC_Src': '00:10:94:02:00:01'}

        return dict10
    def create_spirent_input_dict_PPS_LAG(self,**kwargs):

        self.node = kwargs

        dict10 = {}
        dict10['Spirent_2TAG_AZ'] = {}
        dict10['Spirent_2TAG_ZA'] = {}
        dict10['Spirent_1TAG_AZ'] = {}
        dict10['Spirent_1TAG_ZA'] = {}
        dict10['Spirent_0TAG_AZ'] = {}
        dict10['Spirent_0TAG_ZA'] = {}

        dict10['Spirent_2TAG_AZ']['UC'] = {'Rate_PPS': self.node['active_links'] * 1000,'mac_dst_count' : self.node['active_links'] * 20,'Outer_VLAN_ID': str(self.data['item'] + 100), 'Inner_VLAN_ID': str(self.data['item']),'MAC_Dest':'00:10:94:02:00:01','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_2TAG_ZA']['UC'] = {'Rate_PPS': self.node['active_links'] * 1000,'mac_dst_count' : self.node['active_links'] * 20,'Outer_VLAN_ID': str(self.data['item'] + 100), 'Inner_VLAN_ID': str(self.data['item']),'MAC_Dest':'00:10:94:01:00:01','MAC_Src': '00:10:94:02:00:01'}

        dict10['Spirent_1TAG_AZ']['UC'] = {'Rate_PPS': self.node['active_links'] * 1000,'mac_dst_count' : self.node['active_links'] * 20,'VLAN_ID': str(self.data['item']), 'MAC_Dest':'00:10:94:02:00:01','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_1TAG_ZA']['UC'] = {'Rate_PPS': self.node['active_links'] * 1000,'mac_dst_count' : self.node['active_links'] * 20,'VLAN_ID': str(self.data['item']), 'MAC_Dest':'00:10:94:01:00:01','MAC_Src': '00:10:94:02:00:01'}


        dict10['Spirent_0TAG_AZ']['UC'] = {'Rate_PPS': self.node['active_links'] * 1000,'mac_dst_count' : self.node['active_links'] * 20,'MAC_Dest':'00:10:94:02:00:01','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_0TAG_ZA']['UC'] = {'Rate_PPS': self.node['active_links'] * 1000,'mac_dst_count' : self.node['active_links'] * 20,'MAC_Dest':'00:10:94:01:00:01','MAC_Src': '00:10:94:02:00:01'}


        return dict10
    def create_spirent_input_dict_PPS_FRR(self,**kwargs):

        self.node = kwargs

        dict10 = {}
        dict10['Spirent_2TAG_AZ'] = {}
        dict10['Spirent_2TAG_ZA'] = {}
        dict10['Spirent_1TAG_AZ'] = {}
        dict10['Spirent_1TAG_ZA'] = {}
        dict10['Spirent_0TAG_AZ'] = {}
        dict10['Spirent_0TAG_ZA'] = {}

        dict10['Spirent_2TAG_AZ']['UC'] = {'Rate_PPS': self.node['frr_main_links'] * 1000,'mac_dst_count' : self.node['frr_main_links'] * 20,'Outer_VLAN_ID': str(self.data['item'] + 100), 'Inner_VLAN_ID': str(self.data['item']),'MAC_Dest':'00:10:94:02:00:01','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_2TAG_ZA']['UC'] = {'Rate_PPS': self.node['frr_main_links'] * 1000,'mac_dst_count' : self.node['frr_main_links'] * 20,'Outer_VLAN_ID': str(self.data['item'] + 100), 'Inner_VLAN_ID': str(self.data['item']),'MAC_Dest':'00:10:94:01:00:01','MAC_Src': '00:10:94:02:00:01'}

        dict10['Spirent_1TAG_AZ']['UC'] = {'Rate_PPS': self.node['frr_main_links'] * 1000,'mac_dst_count' : self.node['frr_main_links'] * 20,'VLAN_ID': str(self.data['item']), 'MAC_Dest':'00:10:94:02:00:01','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_1TAG_ZA']['UC'] = {'Rate_PPS': self.node['frr_main_links'] * 1000,'mac_dst_count' : self.node['frr_main_links'] * 20,'VLAN_ID': str(self.data['item']), 'MAC_Dest':'00:10:94:01:00:01','MAC_Src': '00:10:94:02:00:01'}


        dict10['Spirent_0TAG_AZ']['UC'] = {'Rate_PPS': self.node['frr_main_links'] * 1000,'mac_dst_count' : self.node['frr_main_links'] * 20,'MAC_Dest':'00:10:94:02:00:01','MAC_Src': '00:10:94:01:00:01'}
        dict10['Spirent_0TAG_ZA']['UC'] = {'Rate_PPS': self.node['frr_main_links'] * 1000,'mac_dst_count' : self.node['frr_main_links'] * 20,'MAC_Dest':'00:10:94:01:00:01','MAC_Src': '00:10:94:02:00:01'}


        return dict10
    def mep_statistic_accedian(self):
        mep_name = 100000 + self.data['item']
        dict3 = {}
        for node in self.data["site_list"]:          
            if node['login']['device_type'] == 'accedian':
                try:
                    print("**** {} :".format(node['Node_name']))
                    output = node['connect_obj'].send_command("cfm show mep statistics {}".format(node['index']['mep']))
                    print(output)
                    dict3[node['Node_name']] = {}
                    dict3[node['Node_name']]['tx'] = {}
                    dict3[node['Node_name']]['rx'] = {}
                    for dm_sl in ['DM','SL']:
                        for M in ['M','R']:
                            for i in range(7):
                                X = re.findall("{}{} priority {}\s+:\s+\d+".format(dm_sl,M,i), output.replace(",",""))
                                id = '{}{}_P{}'.format(dm_sl,M,i)
                                if '{}{}'.format(dm_sl,M) == 'DMM' or '{}{}'.format(dm_sl,M) == 'SLM':
                                    if int(X[0].split()[-1]) > 0:
                                        dict3[node['Node_name']]['tx'][id] = int(X[0].split()[-1])
                                    if int(X[1].split()[-1]) > 0:
                                        dict3[node['Node_name']]['rx'][id] = int(X[1].split()[-1])
                                else:
                                    if int(X[1].split()[-1]) > 0:
                                        dict3[node['Node_name']]['tx'][id] = int(X[1].split()[-1])
                                    if int(X[0].split()[-1]) > 0:
                                        dict3[node['Node_name']]['rx'][id] = int(X[0].split()[-1])
                except:
                    print("**** something went Wrong Accedian MEP stats could not be checked ")
            else:
                pass                
        return dict3
    def mep_statistic_cisco(self):
        mep_name = 100000 + self.data['item']
        dict4 = {}
        for node in self.data["site_list"]:  
            if node['login']['device_type'] == 'cisco_xr' and 'EP' in node['side']:
                try:
                    output = node['connect_obj'].send_command("show ethernet cfm local meps domain COLT-{} service ALX_NCS_LE-{} verbose".format(self.data['MEG_level'],mep_name))
                    print(output)
                    dict4[node['Node_name']] = {}
                    dict4[node['Node_name']]['tx'] = {}
                    dict4[node['Node_name']]['rx'] = {}
                    for dm_sl in ['DM','SL']:
                        for M in ['M','R']:                
                                X = re.findall("{}{}\s+\w+\s+\w+".format(dm_sl,M), output)
                                #print(X[0].split())
                                id = '{}{}'.format(dm_sl,M)
                                dict4[node['Node_name']]['tx'][id] = int(X[0].split()[1])
                                dict4[node['Node_name']]['rx'][id] = int(X[0].split()[-1])
                except:
                    print("*** something went Wrong, Cisco MEP Stats Could not be checked")
        return dict4
    def create_commands_OLO(self):
        for create_delete in create_delete_list:
            for olo in self.data["OLO_site_list"]:
                with open(file_path + '/templates/create_xc_config_{}_{} copy.j2'.format(olo["side"],create_delete),'r') as f:
                    Temp = f.read()
                    failure_command = Template(Temp).render(**self.data,**olo)
                    file_open = open(file_path + '/commands/XC_command_{}_{}_OLO.txt'.format(olo["Node_name"],create_delete), 'w+')
                    file_open.write(failure_command)
                    file_open.write('\n')
                    file_open.close()
    def push_OLO_config(self):
        for olo in self.data["OLO_site_list"]:
            print("****  Logged in node : {}".format(olo['Node_name']))
            with open(file_path + '/commands/XC_command_{}_create_OLO.txt'.format(olo["Node_name"]),'r') as f:
                f2 = f.readlines()
                output = olo['connect_obj'].send_config_set(f2)
                print(output)
                olo['connect_obj'].commit()
                olo['connect_obj'].exit_config_mode()
    def delete_OLO_config(self):
        for olo in self.data["OLO_site_list"]:
            print("****  Logged in node : {}".format(olo['Node_name']))
            with open(file_path + '/commands/XC_command_{}_delete_OLO.txt'.format(olo["Node_name"]),'r') as f:
                f2 = f.readlines()
                output = olo['connect_obj'].send_config_set(f2)
                print(output)
                olo['connect_obj'].commit()
                olo['connect_obj'].exit_config_mode()               
    def Command_Creation(self):
        self.data['wb'] = load_workbook(filename = file_path + '/QOS.xlsx',read_only=True)                        
        for create_delete in create_delete_list:
            for node in self.data["site_list"]:
                with open(file_path + '/templates/create_xc_config_{}_{} copy.j2'.format(node["side"],create_delete),'r') as f:
                    Temp = f.read()
                    failure_command = Template(Temp).render(**self.data,**node,jinja_func = Jinja_user_func())
                    file_open = open(file_path + '/commands/XC_command_{}_{}.txt'.format(node["Node_name"],create_delete), 'w+')
                    file_open.write(failure_command)
                    # file_open.close()
        self.data['wb'].close()
    def push_config(self):
        for node in self.data["site_list"]:  
            print("**** Logged in node : {}".format(node['Node_name']))
            with open(file_path + '/commands/XC_command_{}_create.txt'.format(node["Node_name"]),'r') as f:
                f2 = f.readlines()
                # output = node['connect_obj'].send_config_set(f2,cmd_verify=False)
                # print(output)    
                if node['login']['device_type'] == 'cisco_xr':
                    output = node['connect_obj'].send_config_set(f2)
                    print(output)
                    node['connect_obj'].commit()
                    node['connect_obj'].exit_config_mode()
                else:
                    output = node['connect_obj'].send_config_set(f2,cmd_verify=False)
                    print(output)
            print("**** Configration completed on {}".format(node['Node_name']))
        print("**** wait for 2 seconds")
        time.sleep(2)
    def push_config_multithread(self):
        def configure(node): 
            print("**** Logged in node {}".format(node['Node_name']))
            with open(file_path + '/commands/XC_command_{}_create.txt'.format(node["Node_name"]),'r') as f:
                f2 = f.readlines()   
            if node['login']['device_type'] == 'cisco_xr':
                node['connect_obj'].send_config_set(f2)
                node['connect_obj'].commit()
                node['connect_obj'].exit_config_mode()
            else:
                node['connect_obj'].send_config_set(f2,cmd_verify=False)
            print("**** Configration completed on {}".format(node['Node_name']))
        with cf.ThreadPoolExecutor(max_workers=5) as ex:
            ex.map(configure, self.data["site_list"])
        time.sleep(3)
    def check_QOS_counters_config(self):
        dict13 = {}
        for node in self.data["site_list"]:
            if 'EP' in node['side']:
                if node['login']['device_type'] == 'cisco_xr':
                    try:
                        if node['port_type'] == 'PL-type':
                            print(f"!\n!!\n**** {node['Node_name']} # show policy-map interface {node['main_interface']} input")
                            output = node['connect_obj'].send_command("show policy-map interface {} input".format(node["main_interface"]))
                        else:
                            print(f"!\n!!\n**** {node['Node_name']} # show policy-map interface {node['main_interface']}.{self.data['item']} input")
                            output = node['connect_obj'].send_command("show policy-map interface {}.{} input".format(node["main_interface"],self.data['item']))
                        print(output)
                        x = re.findall("Total Dropped\s+:\s+\d+", output)
                        # print(x)
                        dict13[node['Node_name']] = int(x[0].split(' ')[-1])
                        if node['port_type'] == 'PL-type':
                            print(f"!\n!!\n**** {node['Node_name']} # show qos interface {node['main_interface']} input")                          
                            output = node['connect_obj'].send_command("show qos interface {} input".format(node["main_interface"]))    
                        else:
                            print(f"!\n!!\n**** {node['Node_name']} # show qos interface {node['main_interface']}.{self.data['item']} input")
                            output = node['connect_obj'].send_command("show qos interface {}.{} input".format(node["main_interface"],self.data['item']))
                        print(output) 
                    except:
                        print("*** something went Wrong, input Policy drops can not be checked")                    
                else:
                    pass
            else:
                pass
        return dict13
    def check_voq_stats(self):

        '''
        try to get the traffic class details through which packets are actually passing
        if bundle, then try to get the stats of all Active interface
        if non LAG, then fire the command directly.
        '''
        dict10 = {}
        for node in self.data["site_list"]:
            if 'EP' in node['side'] or 'IP' in node['side']:
                if node['login']['device_type'] == 'cisco_xr':
                    dict10[f"{node['Node_name']}"] = {}
                    try:
                        if 'Bundle' in node['main_interface']:
                            for interface,status in node['lag_members'].items():
                                if status['state'] == 'Active':
                                    print(f"!\n!!\n**** {node['Node_name']} # show controllers npu stats voq ingress interface {interface} instance all location all")
                                    output = node['connect_obj'].send_command(f"show controllers npu stats voq ingress interface {interface} instance all location all")
                                    print(output)
                                    with open(file_path + '/TEXTFSM/cisco_show_voq.textfsm','r') as f:
                                        re_table = textfsm.TextFSM(f)
                                    fsm_results = re_table.ParseText(output)                                                                
                                    dict_fsm = {f"TC_{items[0]}":{'rx_pkts': items[1]} for items in fsm_results if int(items[1]) != 0 }
                                    dict10[f"{node['Node_name']}"][interface] = dict_fsm

                        else:
                            print(f"!\n!!\n**** {node['Node_name']} # show controllers npu stats voq ingress interface {node['main_interface']} instance all location all")                        
                            output = node['connect_obj'].send_command(f"show controllers npu stats voq ingress interface {node['main_interface']} instance all location all")
                            print(output)
                            with open(file_path + '/TEXTFSM/cisco_show_voq.textfsm','r') as f:
                                re_table = textfsm.TextFSM(f)
                            fsm_results = re_table.ParseText(output)                                                                        
                            dict_fsm = {f"TC_{items[0]}":{'rx_pkts': items[1]} for items in fsm_results if int(items[1]) != 0 }
                            dict10[f"{node['Node_name']}"][node['main_interface']] = dict_fsm
                    except:
                        print("**** something went wrong while getting VOQ stats")
        return dict10
    def check_Mac_table(self):
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr':
                print(f"*** {node['Node_name']} :> show evpn evi mac | include {self.data['item'] + 50000}")
                output = node['connect_obj'].send_command("show evpn evi mac | include {}".format(self.data['item'] + 50000))
                print(output)
                print(f"*** {node['Node_name']} :> show evpn evi neighbor | include {self.data['item'] + 50000}")
                output = node['connect_obj'].send_command("show evpn evi neighbor | include {}".format(self.data['item'] + 50000))
                print(output)
            else:
                pass
    def gather_facts(self):
        self.get_version()
        self.get_lag_status()
        self.get_frr_status()
        self.set_CBS_EBS_for_flatQOS()
        self.get_UNI_NNI_port_speed()
        self.clear_voq_counters()
        self.get_src_tar_evi()
    def get_version(self):
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr':
                output = node['connect_obj'].send_command("show version",use_genie=True)
                node['version'] = output['software_version']
    def get_src_tar_evi(self):
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr':
                node['UID'] = int(node['login']['host'].split('.')[-1])
        if not self.data["ELAN"]:
            for node in self.data["site_list"]:
                if node['login']['device_type'] == 'cisco_xr':
                    for rem_node in self.data["site_list"]:
                        if rem_node['login']['device_type'] == 'cisco_xr':
                            if node['Node_name'] == rem_node['Node_name']:
                                pass
                            else:                  
                                node['src_evi'] = 50000 + self.data['item'] + node['UID']
                                node['tar_evi'] = 50000 + self.data['item'] + rem_node['UID']         

    def clear_voq_counters(self):
        '''
        purpose is to clear the voq counters before the test.
        if bundle then clear the voq counters of all active interfaces
        if not lag, then directly fire the command.
        '''
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr' and 'Bundle' in node['main_interface']:
                for interface,status in node['lag_members'].items():
                    if status['state'] == 'Active':
                        node['connect_obj'].send_command(f"clear controller npu stats voq ingress interface {interface} instance all location 0/0/CPU0")
            elif node['login']['device_type'] == 'cisco_xr' and 'Bundle' not in node['main_interface']:
                node['connect_obj'].send_command(f"clear controller npu stats voq ingress interface {node['main_interface']} instance all location 0/0/CPU0")
            else:
                pass
    def get_UNI_NNI_port_speed(self):
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'accedian':
                try:
                    output = node['connect_obj'].send_command("board show info")
                    x = re.findall("GX|LX|LT", output)
                    if x[0] == 'GX':
                        node['Uni_port_speed'] = 1000000
                        node['Nni_port_speed'] = 1000000
                    else:
                        node['Uni_port_speed'] = 10000000
                        node['Nni_port_speed'] = 10000000
                except:
                    print("**** excepction happened, hence setting Port BW to 10G")
                    node['Uni_port_speed'] = 10000000
                    node['Nni_port_speed'] = 10000000         
    def get_lag_status(self):
        ''' if main interface is LAG/Non LAG, then this function tries to derive
        - how many active links/standby links are there in the LAG
        - if LAG test Can be done
        - what is the distributed service BW per link of LAG.
        - what is LAG total BW or what is main interface BW
        '''
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr' and 'Bundle' in node['main_interface']:
                output = node['connect_obj'].send_command(f"show interface {node['main_interface']}",use_genie=True)
                node['Standby_links'],node['active_links']  = 0,0
                node['Lag_bw'] = output[node['main_interface']]['bandwidth']
                for interfaces,status in output[node['main_interface']]['port_channel']['members'].items():
                    if status['state'] == 'Active':
                        node['active_links'] += 1
                        node['failure_command'] = [f'int {interfaces}','shut']
                        node['repair_command'] = [f'int {interfaces}','no shut'] 
                    elif status['state'] == 'Standby':
                        node['Standby_links'] += 1
                if node['active_links'] >= 2 or (node['active_links'] == 1 and node['Standby_links'] == 1):
                    node['Lag_test_eligible'] = True
                else:
                    node['Lag_test_eligible'] = False
                node['main_interface_bw'] = node['Lag_bw'] // node['active_links']
                node['distributed_service_BW'] = self.data['service_BW'] // node['active_links']
                node['lag_members'] = output[node['main_interface']]['port_channel']['members']
            elif node['login']['device_type'] == 'cisco_xr' and 'Bundle' not in node['main_interface']:
                output = node['connect_obj'].send_command(f"show interface {node['main_interface']}",use_genie=True)
                node['distributed_service_BW'] = self.data['service_BW']
                node['main_interface_bw'] = output[node['main_interface']]['bandwidth']
    def delete_config(self):
        for node in self.data["site_list"]:
            with open(file_path + '/commands/XC_command_{}_delete.txt'.format(node["Node_name"]),'r') as f:
                f2 = f.readlines()
                if node['login']['device_type'] == 'cisco_xr':
                    output = node['connect_obj'].send_config_set(f2)
                    print(output)
                    node['connect_obj'].commit()
                    node['connect_obj'].exit_config_mode()
                else:
                    output = node['connect_obj'].send_config_set(f2,cmd_verify=False)
                    print(output)
        if self.data['SRTE']:
            for P_C_G_D in ['delete']:
                with open(file_path + '/commands/htool-{}.txt'.format(format(P_C_G_D),'r')) as f:
                    for f2 in f:
                        print(f2)
                        output = self.data['Htool']['connect_obj'].send_command(f2)
                        print(output)
                    print("****  {} completed ".format(P_C_G_D))
                    print("**** wait for 2 seconds")
                    time.sleep(2)
        for file_name in listdir(f"{file_path}/commands/"):    
            if file_name.endswith('.txt'):    
                os.remove(f"{file_path}/commands/" + file_name)            
    def parse_accedian(self):
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr':
                node['UID'] = int(node['login']['host'].split('.')[-1])
            else:                
                print("****  Logged in node : {}".format(node['Node_name']))
                node['index'] = {}
                if node['Protected'] == 'YES':
                    node['out_port'] = 'LAG-{}'.format(node['Nni_port'] // 2 + 1)
                else:
                    node['out_port'] = 'PORT-{}'.format(node['Nni_port'])                                
                for mep_meg_dmm_slm in mep_meg_dmm_slm_list:
                    output = node['connect_obj'].send_command('cfm show {} configuration'.format(mep_meg_dmm_slm))
                    template = open(file_path + '/TEXTFSM/accedian_show_{}_index.textfsm'.format(mep_meg_dmm_slm))
                    re_table = textfsm.TextFSM(template)
                    fsm_results = re_table.ParseText(output)
                    # if mep_meg_dmm_slm == 'meg':
                    #     node['index']['del_meg'] = 1
                    #     for rows in fsm_results:
                    #         if rows[1] == 'LEXXX-{}'.format(100000 + self.data['item']):
                    #             node['index']['del_meg'] = int(rows[0])
                    if len(fsm_results) == 0:
                        node['index'][mep_meg_dmm_slm] = 1
                    else:                   
                        node['index'][mep_meg_dmm_slm] = int(fsm_results[-1][0]) + 1
                    if mep_meg_dmm_slm == 'meg':
                        node['index']['del_meg'] = node['index'][mep_meg_dmm_slm]
                print("****  persing completed on {}".format(node['Node_name']))
                print(node['index'])

        return node['index']
    def Validate_ccm(self):
        ''' Purpose:
        - to Check CCM status on Accedian
        - to Check ccm status on NCS if not ELAN.
        '''
        test_result = {}
        for node in self.data["site_list"]:
            mep_name = 100000 + self.data['item']
            if 'EP' in node['side']:
                if node['login']['device_type'] == 'accedian':
                    print(f"**** {node['Node_name']}",end=' : ')
                    output = node['connect_obj'].send_command("cfm show mep status LEXXX-{}|{}|{}".format(mep_name,self.data['MEG_level'],node['Remote_MEP']))
                    if len(re.findall("Inactive", output)) == 14:
                        print("ccm is UP")
                        test_result[node['Node_name']] = 'pass'
                    else:
                        print("ccm did not came Up")
                        test_result[node['Node_name']] = 'fail'      
                else:
                    if self.data["ELAN"]:
                        print("**** {}".format(node['Node_name']),end=' : ')
                        print("ccm not applicable for ELAN on NCS")
                        test_result[node['Node_name']] = 'NA'
                    else:
                        print("**** {}".format(node['Node_name']),end=' : ')
                        output = node['connect_obj'].send_command("show ethernet cfm services domain COLT-{} service ALX_NCS_LE-{}".format(self.data['MEG_level'],mep_name))
                        if len(re.findall("all operational, no errors", output)) == 2:
                            print("ccm is UP")
                            test_result[node['Node_name']] = 'pass'
                        else:
                            print("ccm did not came Up")
                            test_result[node['Node_name']] = 'fail'
        return test_result
    def Y1564_test(self):
        list1 = []
        test_result = {}
        self.data["loop_ext_int"] = 'internal'
        mep_name = 100000 + self.data['item']
        for node in self.data["site_list"]:
            if 'EP' in node['side']:
                list1.append(node['login']['device_type'])
            else:
                pass
        if list1 == ['accedian','accedian']:
            for node in self.data["site_list"]:
                if node['login']['device_type'] == 'accedian':
                    output = node['connect_obj'].send_command("cfm show mep database LEXXX-{}|{}|{}".format(mep_name,self.data['MEG_level'],node['Remote_MEP']))
                    node['remote_mac'] = re.findall("\w\w[:]\w\w[:]\w\w[:]\w\w[:]\w\w[:]\w\w", output)[0]
                    if node['Protected'] == 'YES':
                        output = node['connect_obj'].send_command("port show status PORT-{}".format(node['Nni_port']))
                        if re.findall("Down|Up", output)[0] == 'Down':
                            node['Nni_port'] = node['Nni_port'] + 1
                        else:
                            pass
                    else:
                        pass
                    node['packet_type'] = 'l2-accedian'
                    for create_delete in create_delete_list:
                        with open(file_path + '/templates/Accedian_{}_{}_Y1564.j2'.format(node["side"],create_delete),'r') as f:
                            Temp = f.read()
                            failure_command = Template(Temp).render(**self.data,**node)
                            file_open = open(file_path + '/commands/Accedian_{}_{}_Y1564.txt'.format(node["Node_name"],create_delete), 'w+')
                            file_open.write(failure_command)
                            file_open.write('\n')
                            file_open.close()
                            print("**** {} templating done on node {} ".format(create_delete,node['Node_name']))
                            with open(file_path + '/commands/Accedian_{}_{}_Y1564.txt'.format(node["Node_name"],create_delete),'r') as f:
                                f2 = f.readlines()
                                output = node['connect_obj'].send_config_set(f2,cmd_verify=False)
                                print(output)
                            if create_delete == "create":
                                time.sleep(10)
                                output = node['connect_obj'].send_command("Y1564 show activation Y1564-LE-{}".format(mep_name))
                                print(output)
                                x = re.findall("FAILED|PROGRESS", output)
                                if x[0] == 'FAILED':
                                    test_result[node['Node_name']] = 'fail'
                                else:
                                    time_to_wait = (self.data["config_test"]*4) + (self.data["performance_test"]*60) + 30
                                    print("***  Hold your breathe for {} seconds".format(time_to_wait))
                                    time.sleep(time_to_wait)
                                    output = node['connect_obj'].send_command("Y1564 show activation Y1564-LE-{}".format(mep_name))
                                    print(output)                             
                                    x = re.findall("PASSED|FAILED|PROGRESS", output)
                                    if x[0] == 'PASSED':
                                        test_result[node['Node_name']] = 'pass'
                                    elif x[0] == 'FAILED':
                                        test_result[node['Node_name']] = 'fail'
                                    else:
                                        test_result[node['Node_name']] = 'something Wrong,still in progress'
                                    
                                                                
        elif list1 == ['cisco_xr','cisco_xr']:
            if self.data['ELAN']:
                pass
            else:                
                Loop_list = ['L2'] # valid lists are ['L1','L2']
                ''' perse the Jinja files to Create the commands for Looping remote end traffic generation at Local End'''
                for node in self.data["site_list"]:
                    node['loop_ID'] = 1
                    node['delete_L1_loop'] = True
                    ''' try to find if Loopback internal is given at the other end.'''
                    output = node['connect_obj'].send_command("show run int {}".format(node['main_interface']))
                    x = re.findall("loopback internal",output)
                    if len(x) == 1:
                        node['delete_L1_loop'] = False
                    ''' Y1564 remote mac is = local mac, in case of L1 Loop'''
                    for looptype in Loop_list:
                        if looptype == 'L1':
                            node['remote_mac'] = node['connect_obj'].send_command("show interface {}".format(node['main_interface']),use_genie=True)[node['main_interface']]['mac_address']
                        else:
                            node['remote_mac'] = '0022.0022.0022'
                        ''' Create the templates '''
                        for pro_loop in ['profile','loop']:
                            for create_delete in create_delete_list:
                                with open(file_path + '/templates/Cisco_{}_{}_{}_Y1564.j2'.format(looptype,pro_loop,create_delete),'r') as f:
                                    Temp = f.read()       
                                    command = Template(Temp).render(**self.data,**node)
                                    file_open = open(file_path + '/commands/Cisco_{}_{}_{}_{}_Y1564.txt'.format(node["Node_name"],looptype,pro_loop,create_delete), 'w+')
                                    file_open.write(command)
                                    file_open.write('\n')
                                    file_open.close()
                                    print("*** {} {} {} commands are templated for {}".format(looptype,pro_loop,create_delete,node['Node_name']))
                ''' perse completed and txt files are generated '''
                ''' now loop over 2 times on the same site list, and if local node == remote node, pass that case '''
                for node in self.data["site_list"]:
                    for remote_node in self.data["site_list"]:
                        if node['Node_name'] == remote_node['Node_name']:
                            pass
                        else:
                            ## added and node["Node_name"] == 'AR15'
                            if node['version'] == '7.1.2' and node["Node_name"] == 'AR15':
                                ''' if node is capable, perform loop on the other end & parse the loop ID '''
                                test_result[node["Node_name"]] = {}
                                for looptype in Loop_list:
                                    if 'Bundle' in remote_node['main_interface'] and looptype == 'L1':
                                        test_result[node["Node_name"]][looptype] = 'NA_LAG'
                                        pass
                                    else:
                                        for create_delete in create_delete_list:
                                            if create_delete == 'delete' and looptype == 'L1' and not remote_node['delete_L1_loop']:
                                                print(remote_node['delete_L1_loop'])
                                                pass
                                            else:
                                                with open(file_path + '/commands/Cisco_{}_{}_loop_{}_Y1564.txt'.format(remote_node['Node_name'],looptype,create_delete),'r') as f:
                                                    f2 = f.readlines()
                                                    output = remote_node['connect_obj'].send_config_set(f2)
                                                    remote_node['connect_obj'].commit()
                                                    remote_node['connect_obj'].exit_config_mode()
                                                    print(output)
                                                    ''' perse the loop id if L2 Loop is configured '''
                                                    if create_delete == 'create' and looptype == 'L2':
                                                        output = remote_node['connect_obj'].send_command("show ethernet loopback active | in ID")
                                                        remote_node['loop_ID'] = int(re.split("\s", output)[-1])
                                                        if remote_node['loop_ID'] != 1:
                                                            ''' if Loop id not equals 1 then create the delete loop commands with the new ID '''
                                                            with open(file_path + '/templates/Cisco_L2_loop_delete_Y1564.j2','r') as f:
                                                                Temp = f.read()
                                                                failure_command = Template(Temp).render(**self.data,**remote_node)
                                                                file_open = open(file_path + '/commands/Cisco_{}_{}_loop_delete_Y1564.txt'.format(remote_node["Node_name"],looptype), 'w+')
                                                                file_open.write(failure_command)
                                                                file_open.write('\n')
                                                                file_open.close()                                    
                                            ''' configure 1564 profile at local end & start the Test'''
                                            with open(file_path + '/commands/Cisco_{}_{}_profile_{}_Y1564.txt'.format(node["Node_name"],looptype,create_delete),'r') as f:
                                                f2 = f.readlines()
                                                output = node['connect_obj'].send_config_set(f2)
                                                node['connect_obj'].commit()
                                                node['connect_obj'].exit_config_mode()
                                                print(output)
                                            ''' if loop is under create then check if the test is not aborted, else wait for performance test duration & parse Frame loss'''
                                            if create_delete == 'create':
                                                time.sleep(2)
                                                if node['port_type'] == 'PL-type':
                                                    command = f"show ethernet service-activation-test {node['main_interface']}"
                                                else:
                                                    command = f"show ethernet service-activation-test {node['main_interface']}.{self.data['item']}"
                                                output = node['connect_obj'].send_command(command)
                                                x = re.findall("aborted",output)
                                                if len(x) == 1:
                                                    test_result[node["Node_name"]][looptype] = 'aborted'
                                                    print(output)
                                                else:
                                                    print("*** {} loop, Hold your breath for {} seconds, packets are on the wire".format(looptype,self.data['performance_test']*60 + 20))                                                        
                                                    ''' if duration is more than 2 Mnts then disconnect the SSH connection & connect after the specified duration'''
                                                    if self.data['performance_test'] > 2:
                                                        self.disconnect_nodes()
                                                        for i in tqdm(range(self.data['performance_test']*60 + 20),ncols=100,desc='Y1564 Progress'):
                                                            time.sleep(1)
                                                        self.connect_nodes()
                                                    else:
                                                        for i in tqdm(range(self.data['performance_test']*60 + 20),ncols=100,desc='Y1564 Progress'):
                                                            time.sleep(1)
                                                    with open(file_path + '/netconf_filters/filter_ethernet_service_test_state.j2') as f:
                                                        xml_structure = Template(f.read()).render(**self.data,**node)                                                  
                                                    reply_xml = node['net_conf_obj'].get(xml_structure)
                                                    reply_dict = xmltodict.parse(reply_xml.xml)["rpc-reply"]["data"]["service-activation-test"]["test-statuses"]["test-status"]["phase"]["results"]
                                                    output_dict = json.loads(json.dumps(reply_dict))
                                                    # pprint(output_dict)                                    
                                                    output = node['connect_obj'].send_command("show ethernet service-activation-test")
                                                    print(output)
                                                    ''' parse the Output and see FL == 0 '''
                                                    x = re.findall("FL: \d+",output)
                                                    try: 
                                                        if int(x[0].split()[1]) == 0 and int(x[1].split()[1]) == 0 :
                                                            test_result[node["Node_name"]][looptype]= 'pass'
                                                            test_result[node["Node_name"]][f"{looptype}_CIR_TX"] = output_dict['cir']['tx-packets']
                                                            test_result[node["Node_name"]][f"{looptype}_CIR_FL"] = output_dict['cir']['frames-lost']
                                                            test_result[node["Node_name"]][f"{looptype}_EIR_TX"] = output_dict['eir']['tx-packets']
                                                            test_result[node["Node_name"]][f"{looptype}_EIR_FL"] = output_dict['eir']['frames-lost']
                                                        else:
                                                            test_result[node["Node_name"]][looptype] = 'fail'
                                                            test_result[node["Node_name"]][f"{looptype}_CIR_TX"] = output_dict['cir']['tx-packets']
                                                            test_result[node["Node_name"]][f"{looptype}_CIR_FL"] = output_dict['cir']['frames-lost']
                                                            test_result[node["Node_name"]][f"{looptype}_EIR_TX"] = output_dict['eir']['tx-packets']
                                                            test_result[node["Node_name"]][f"{looptype}_EIR_FL"] = output_dict['eir']['frames-lost']
                                                    except:
                                                        test_result[node["Node_name"]][looptype] = 'exception'
                            else:
                                test_result[node["Node_name"]] = 'Not_supported'
            return test_result                                                      

        elif list1 == ['cisco_xr','accedian'] or list1 == ['accedian','cisco_xr']:
            ''' identify the other end to decide the loop List'''
            for node in self.data["site_list"]:
                if node['login']['device_type'] == 'cisco_xr' and 'EP' in node['side']:
                    if "Bundle" in node['main_interface']:
                        Loop_list = ['L2']
                    else:
                        Loop_list = ['L1','L2']
            ''' prepare the templates '''
            for node in self.data["site_list"]:
                if node['login']['device_type'] == 'accedian' and 'EP' in node['side']:
                    test_result[node["Node_name"]] = {}
                    if node['Protected'] == 'YES':
                        output = node['connect_obj'].send_command("port show status PORT-{}".format(node['Nni_port']))
                        if re.findall("Down|Up", output)[0] == 'Down':
                            node['Nni_port'] = node['Nni_port'] + 1
                    output = node['connect_obj'].send_command("port show configuration PORT-{}".format(node['Nni_port']))
                    node['remote_mac'] = re.findall("\w\w[:]\w\w[:]\w\w[:]\w\w[:]\w\w[:]\w\w", output)[0]                       
                    node['packet_type'] = 'l2-generic'                    
                    for create_delete in create_delete_list:
                        for looptype in Loop_list:
                            if looptype == 'L2':
                                node['remote_mac'] = '00:22:00:22:00:22'
                            with open(file_path + '/templates/Accedian_{}_{}_Y1564.j2'.format(node["side"],create_delete),'r') as f:
                                failure_command = Template(f.read()).render(**self.data,**node)
                                file_open = open(file_path + '/commands/Accedian_{}_{}_{}_Y1564.txt'.format(node["Node_name"],looptype,create_delete), 'w+')
                                file_open.write(failure_command)
                                file_open.close()
                                print("*** Loop {} commands are templated for {}".format(create_delete,node['Node_name']))
                elif node['login']['device_type'] == 'cisco_xr' and 'EP' in node['side']:
                    node['loop_ID'] = 1
                    for create_delete in create_delete_list:
                        for looptype in Loop_list:
                            with open(file_path + '/templates/Cisco_{}_loop_{}_Y1564.j2'.format(looptype,create_delete),'r') as f:
                                failure_command = Template(f.read()).render(**self.data,**node)
                                file_open = open(file_path + '/commands/Cisco_{}_{}_{}_Y1564.txt'.format(node["Node_name"],looptype,create_delete), 'w+')
                                file_open.write(failure_command)
                                file_open.close()
                                print("*** {} Loop {} commands are templated for {}".format(looptype,create_delete,node['Node_name']))
            for looptype in Loop_list:
                for create_delete in create_delete_list:
                    for node in self.data["site_list"]:
                        if node['login']['device_type'] == 'cisco_xr' and 'EP' in node['side']:
                            with open(file_path + '/commands/Cisco_{}_{}_{}_Y1564.txt'.format(node["Node_name"],looptype,create_delete),'r') as f:
                                output = node['connect_obj'].send_config_set(f.readlines())
                                node['connect_obj'].commit()
                                node['connect_obj'].exit_config_mode()
                                print(output)
                                if looptype == 'L2' and create_delete == 'create':
                                    output = node['connect_obj'].send_command("show ethernet loopback active | in ID")
                                    node['loop_ID'] = re.split("\s", output)[-1]
                                    with open(file_path + '/templates/Cisco_L2_loop_delete_Y1564.j2','r') as f:
                                        failure_command = Template(f.read()).render(**self.data,**node)
                                        file_open = open(file_path + '/commands/Cisco_{}_L2_delete_Y1564.txt'.format(node["Node_name"]), 'w+')
                                        file_open.write(failure_command)
                                        file_open.close()
                        elif node['login']['device_type'] == 'accedian' and 'EP' in node['side']:
                            with open(file_path + '/commands/Accedian_{}_{}_{}_Y1564.txt'.format(node["Node_name"],looptype,create_delete),'r') as f:
                                output = node['connect_obj'].send_config_set(f.readlines(),cmd_verify=False)
                                print(output)
                                if create_delete == "create":
                                    time.sleep(10)
                                    output = node['connect_obj'].send_command("Y1564 show activation Y1564-LE-{}".format(mep_name))
                                    print(output)
                                    x = re.findall("FAILED|PROGRESS", output)
                                    if x[0] == 'FAILED':
                                        test_result[node["Node_name"]] = 'fail'
                                    else:
                                        time_to_wait = (self.data["config_test"]*4) + (self.data["performance_test"]*60) + 20
                                        print("***  Hold your breathe for {} seconds".format(time_to_wait))
                                        for i in tqdm(range(time_to_wait),ncols=100,desc='Y1564 Progress'):
                                            time.sleep(1)
                                        output = node['connect_obj'].send_command("Y1564 show activation Y1564-LE-{}".format(mep_name))
                                        print(output)                             
                                        x = re.findall("PASSED|FAILED|PROGRESS", output)
                                        if x[0] == 'PASSED':
                                            test_result[node["Node_name"]][looptype] = 'pass'
                                        elif x[0] == 'FAILED':
                                            test_result[node["Node_name"]][looptype] = 'fail'
                                        else:
                                            test_result[node["Node_name"]][looptype] = 'something Wrong,still in progress'
                                        pprint(test_result)            
            '''
            Template the configurations
            set L1 Loop on Accedian
            Start Y1564 on Cisco
            Set L2 on Accdian NNI Port
            Start Y1564 on cisco
            '''
            for node in self.data["site_list"]:
                if node['login']['device_type'] == 'cisco_xr' and 'EP' in node['side']:
                    if node['version'] == '7.1.2':
                        test_result[node["Node_name"]] = {}
                        for node in self.data["site_list"]:
                            if node['login']['device_type'] == 'accedian' and 'EP' in node['side']:
                                for looptype in ['L1','L2']:
                                    for create_delete in ['create','delete']:
                                        with open(file_path + f"/templates/Accedian_{looptype}_loop_{create_delete}_Y1564.j2",'r') as f:
                                            templated_command = Template(f.read()).render(**self.data,**node)
                                            file_open = open(file_path + f"/commands/Accedian_{node['Node_name']}_{looptype}_loop_{create_delete}_Y1564.txt", 'w+')
                                            file_open.write(templated_command)
                                            file_open.close()
                                            print(f"**** {looptype} loop {create_delete} commands templated for {node['Node_name']}")
                        for node in self.data["site_list"]:
                            if node['login']['device_type'] == 'cisco_xr' and 'EP' in node['side']:
                                for looptype in ['L1','L2']:
                                    if looptype == 'L1':
                                        node['remote_mac'] = node['connect_obj'].send_command("show interface {}".format(node['main_interface']),use_genie=True)[node['main_interface']]['mac_address']
                                    else:
                                        node['remote_mac'] = '0022.0022.0022'    
                                    for create_delete in ['create','delete']:
                                        with open(file_path + f"/templates/Cisco_{looptype}_profile_{create_delete}_Y1564.j2",'r') as f:
                                            templated_command = Template(f.read()).render(**self.data,**node)
                                            file_open = open(file_path + f"/commands/Cisco_{node['Node_name']}_{looptype}_profile_{create_delete}_Y1564.txt", 'w+')
                                            file_open.write(templated_command)
                                            file_open.close()
                                            print(f"**** {looptype} profile {create_delete} commands templated for {node['Node_name']}")
                        for looptype in ['L1','L2']:
                            for create_delete in ['create','delete']:
                                for node in self.data["site_list"]:
                                    if node['login']['device_type'] == 'accedian' and 'EP' in node['side']:
                                        with open(file_path + '/commands/Accedian_{}_{}_loop_{}_Y1564.txt'.format(node["Node_name"],looptype,create_delete),'r') as f:
                                            output = node['connect_obj'].send_config_set(f.readlines(),cmd_verify=False)
                                            print(output)                        
                                for node in self.data["site_list"]:
                                    if node['login']['device_type'] == 'cisco_xr' and 'EP' in node['side']:
                                        with open(file_path + '/commands/Cisco_{}_{}_profile_{}_Y1564.txt'.format(node["Node_name"],looptype,create_delete),'r') as f:
                                            output = node['connect_obj'].send_config_set(f.readlines())
                                            node['connect_obj'].commit()
                                            node['connect_obj'].exit_config_mode()
                                            print(output)
                                        if create_delete == 'create':
                                            time.sleep(2)
                                            command = f"show ethernet service-activation-test {node['main_interface']}.{self.data['item']}"
                                            output = node['connect_obj'].send_command(command)
                                            x = re.findall("aborted",output)
                                            if len(x) == 1:
                                                test_result[node["Node_name"]][looptype] = 'aborted'
                                                print(output)
                                            else:
                                                print("*** {} loop, Hold your breath for {} seconds, packets are on the wire".format(looptype,self.data['performance_test']*60 + 20))
                                                ''' if duration is more than 2 Mnts then disconnect the SSH connection & connect after the specified duration'''
                                                if self.data['performance_test'] > 2:
                                                    self.disconnect_nodes()
                                                    for i in tqdm(range(self.data['performance_test']*60 + 20),ncols=100,desc='Y1564 Progress'):
                                                        time.sleep(1)
                                                    self.connect_nodes()
                                                else:
                                                    for i in tqdm(range(self.data['performance_test']*60 + 20),ncols=100,desc='Y1564 Progress'):
                                                        time.sleep(1)
                                                with open(file_path + '/netconf_filters/filter_ethernet_service_test_state.j2') as f:
                                                    xml_structure = Template(f.read()).render(**self.data,**node)                                                  
                                                reply_xml = node['net_conf_obj'].get(xml_structure)
                                                reply_dict = xmltodict.parse(reply_xml.xml)["rpc-reply"]["data"]["service-activation-test"]["test-statuses"]["test-status"]["phase"]["results"]
                                                output_dict = json.loads(json.dumps(reply_dict))
                                                # pprint(output_dict)                                    
                                                output = node['connect_obj'].send_command("show ethernet service-activation-test")
                                                print(output)
                                                ''' parse the Output and see FL == 0 '''
                                                x = re.findall("FL: \d+",output)
                                                try: 
                                                    if int(x[0].split()[1]) == 0 and int(x[1].split()[1]) == 0 :
                                                        test_result[node["Node_name"]][looptype]= 'pass'
                                                        test_result[node["Node_name"]][f"{looptype}_CIR_TX"] = output_dict['cir']['tx-packets']
                                                        test_result[node["Node_name"]][f"{looptype}_CIR_FL"] = output_dict['cir']['frames-lost']
                                                        test_result[node["Node_name"]][f"{looptype}_EIR_TX"] = output_dict['eir']['tx-packets']
                                                        test_result[node["Node_name"]][f"{looptype}_EIR_FL"] = output_dict['eir']['frames-lost']
                                                    else:
                                                        test_result[node["Node_name"]][looptype] = 'fail'
                                                        test_result[node["Node_name"]][f"{looptype}_CIR_TX"] = output_dict['cir']['tx-packets']
                                                        test_result[node["Node_name"]][f"{looptype}_CIR_FL"] = output_dict['cir']['frames-lost']
                                                        test_result[node["Node_name"]][f"{looptype}_EIR_TX"] = output_dict['eir']['tx-packets']
                                                        test_result[node["Node_name"]][f"{looptype}_EIR_FL"] = output_dict['eir']['frames-lost']
                                                except:
                                                    test_result[node["Node_name"]][looptype] = 'exception'
                    else:
                        test_result[node["Node_name"]] = 'Not_suported'



        else:
            pass
        return test_result
    def get_frr_status(self):
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr':
                for node2 in self.data["site_list"]:
                    if node2['login']['device_type'] == 'cisco_xr':
                        if node['login']['host'] == node2['login']['host']:
                            pass
                        else:
                            node['remote_rid'] = node2['login']['host']
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr':
                output = node['connect_obj'].send_command(f"show route ipv4 {node['remote_rid']}/32 | include , via")
                # print(output)
                x = re.findall("HundredGigE\d/\d/\d/\d", output)
                y = re.findall("Protected| Backup",output)
                if len(x) == 1:
                    node['frr_main_links'] = 1
                    node['frr_primary'] = x[0]
                    node['frr_test_eligible'] = False
                else:
                    node['frr_main_links'] = 0
                    node['frr_test_eligible'] = True                        
                    for l,n in zip(x,y):
                        if "Protected" in n:
                            node['frr_primary'] = l
                            node['frr_main_links'] = node['frr_main_links'] + 1
                            
            # pprint(node)                           
    def SRTE_Config(self):
        try:
            if self.data["SRTE"]:
                Node_count = 0
                for node in self.data["site_list"]:
                    if node['login']['device_type'] == 'cisco_xr':
                        Node_count = Node_count + 1
                if Node_count == 2:
                        self.data['Htool']['src1'] = self.data["site_list"][0]['login']['host']
                        self.data['Htool']['src2'] = self.data["site_list"][0]['login']['host']
                        self.data['Htool']['dst1'] = self.data["site_list"][1]['login']['host']
                        self.data['Htool']['dst2'] = self.data["site_list"][1]['login']['host']
                for P_C_G_D in ['preview','get','create','delete']:
                    with open(file_path + '/templates/htool-{}.j2'.format(P_C_G_D),'r') as f:
                        Temp = f.read()
                        failure_command = Template(Temp).render(**self.data['Htool'])
                        file_open = open(file_path + '/commands/htool-{}.txt'.format(P_C_G_D), 'w+')
                        file_open.write(failure_command)
                        file_open.write('\n')
                        file_open.close()
                for P_C_G_D in ['preview','create','get']:
                    with open(file_path + '/commands/htool-{}.txt'.format(format(P_C_G_D),'r')) as f:
                        loop = 0
                        for f2 in f:
                            print(f2)
                            output = self.data['Htool']['connect_obj'].send_command(f2)
                            if P_C_G_D == 'get':
                                x = re.findall("PW name:\\s\\S+", output)
                                self.data["site_list"][loop]['pw_class'] = x[1].split()[2]
                            print(output)
                            loop = loop + 1
                        print("****  {} completed ".format(P_C_G_D))
                        print("**** wait for 2 seconds")
                        time.sleep(2)
        except:
            print("*** some error is there")        
    def get_netconf_XC_status(self):
        result = {}
        if not self.data['ELAN']:
            ### template the jinja file with EVI ID(item) & store the xml file in variable named xml_structure.
            with open(file_path + '/netconf_filters/filter_Xconnect_state.j2') as f:
                xml_structure = Template(f.read()).render(**self.data)
            ### now if device is cisco_xr, then send the xml to get xconnect status via netconf connection object
            for node in self.data["site_list"]:
                if node['login']['device_type'] == 'cisco_xr':
                    try:                   
                        reply_xml = node['net_conf_obj'].get(xml_structure)
                        ### convert the reply xml into dictionary
                        reply_dict = xmltodict.parse(reply_xml.xml)["rpc-reply"]["data"]["l2vpnv2"]["nodes"]["node"]["xconnects"]["xconnect"]
                        ### convert ordered dict to a regular dictionary, for easy view
                        output_dict = json.loads(json.dumps(reply_dict))
                        ### print the Regular Dictionary
                        # pprint(output_dict)
                        ### do validation of the status of AC and Pseudowire
                        if 'down' not in output_dict['segment1']['attachment-circuit']['state'] and 'down' not in output_dict['segment2']['pseudo-wire']['state']:
                            print(f"**** service is UP at {node['Node_name']}")
                            result[node['Node_name']] = 'pass'
                        else:
                            print(f"**** service is down at {node['Node_name']}")
                            result[node['Node_name']] = 'fail'
                    except:
                        print("*** Something went wrong")
        return result
    def netconf_set_random_MTU(self,flag):
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr':
                try:
                    if flag == 'random':
                        node['rand_mtu'] = random.randrange(2000,9186,2)
                    else:
                        node['rand_mtu'] = 9186                    
                    with open(file_path + '/netconf_filters/config_MTU_interface.j2') as f:
                        xml_structure = Template(f.read()).render(**self.data,**node)
                    # print(xml_structure)       
                    node['net_conf_obj'].edit_config(xml_structure,target='candidate',default_operation='merge')
                    node['net_conf_obj'].commit()
                    print(f"**** {node['rand_mtu']} MTU confgiured on {node['Node_name']}")
                except:
                    print("**** something went Wrong while setting MTU")
        time.sleep(3)
    def get_netconf_BGP_status(self):
        result = {}
        with open(file_path + '/netconf_filters/filter_bgp_neighbor.j2') as f:
            xml_structure = Template(f.read()).render(**self.data)
        ### now if device is cisco_xr, then send the xml to get xconnect status via netconf connection object
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr':
                try:                   
                    reply_xml = node['net_conf_obj'].get(xml_structure)
                    ### convert the reply xml into dictionary
                    reply_dict = xmltodict.parse(reply_xml.xml)["rpc-reply"]["data"]['bgp']['instances']['instance']['instance-active']['default-vrf']['neighbors']['neighbor']
                    ### convert ordered dict to a regular dictionary, for easy view
                    output_dict = json.loads(json.dumps(reply_dict))
                    ### print the Regular Dictionary
                    # pprint(output_dict)
                    if type(output_dict) is list:
                        node['bgp'] = output_dict
                    else:
                        node['bgp'] = []
                        node['bgp'].append(output_dict)
                except:
                    print("*** Something went wrong")
        return result
    def netconf_shut_bgp_neighbor(self,flag):
        for node in self.data["site_list"]:
            if node['login']['device_type'] == 'cisco_xr':
                ###  set the bgp action flag(true or false)
                node['bgp_action'] = flag
                ### render the jinja file to get the edit_config xml structure.
                with open(file_path + '/netconf_filters/shut_bgp_neighbor.j2') as f:
                    xml_structure = Template(f.read()).render(**self.data,**node)
                try:                   
                    ### Merge xml_structure in the Candiate DB store & verify rpc_reply contains <ok/> or not.
                    response = node['net_conf_obj'].edit_config(xml_structure,target='candidate')
                    if '<ok/>' in response.xml:
                        print(f"**** Candidate config merged Correctly on {node['Node_name']}")
                    ### commit Candidate DB store as Running config & verify rpc_reply contains <ok/> or not.
                    response = node['net_conf_obj'].commit()
                    if '<ok/>' in response.xml:                    
                        print(f"**** bgp neighbor shut = {flag} completed on {node['Node_name']}")
                    time.sleep(2)
                except:
                    print("something went Wrong")
        print("**** wait for 10 seconds")
        time.sleep(10)
    def set_CBS_EBS_for_flatQOS(self):
        if self.data["Flat_QOS"]:
            ''' load the QOS xlsx file in readonly mode '''
            wb = load_workbook(filename = file_path +'/QOS.xlsx',read_only=True)
            ''' identify main interface BW, like 1G,10G,100G or 25G, if Uni is Bundle then its handled in Other Function'''    
            for node in self.data["site_list"]:
                if node['login']['device_type'] == 'cisco_xr':
                    ''' Now try to find the CBS EBS Value from sheet for Flat QOS and HQOS'''
                    for row in wb[self.data["QOS_type"]].rows:                        
                        if node['main_interface_bw'] == row[0].value and node['distributed_service_BW'] == row[1].value:
                            node['cbs'],node['ebs'] = row[2].value,row[3].value
                        elif node['main_interface_bw'] == row[0].value and node['distributed_service_BW'] > row[1].value and node['distributed_service_BW'] < row[4].value:
                            node['cbs'],node['ebs'] = row[2].value,row[3].value
            wb.close()
    def Y1564_test_external(self,NCS_node,acc_node):
        ''' set the loop type parameter'''
        test_result = {}
        for node in self.data["site_list"]:
            if node['Node_name'] == NCS_node:
                if node['version'] == '7.1.2' and not self.data['ELAN']:
                    test_result[node["Node_name"]] = {}
                    self.data["loop_ext_int"] = 'external'
                    ''' template the files'''
                    Loop_list = ['L1'] # valid lists are ['L1','L2']
                    ''' perse the Jinja files to Create the commands for Looping remote end traffic generation at Local End'''
                    for node in self.data["site_list"]:
                        if node['Node_name'] == NCS_node or node['Node_name'] == acc_node:
                            if node['login']['device_type'] == 'cisco_xr':
                                ''' preapre the Cisco templates '''
                                for create_delete in ['create','delete']:
                                    for looptype in Loop_list:
                                        if looptype == 'L1':
                                            node['remote_mac'] = node['connect_obj'].send_command("show interface {}".format(node['main_interface']),use_genie=True)[node['main_interface']]['mac_address']
                                        else:
                                            node['remote_mac'] = '0022.0022.0022' 
                                        with open(file_path + f"/templates/Cisco_{looptype}_profile_{create_delete}_Y1564.j2",'r') as f:
                                            templated_command = Template(f.read()).render(**self.data,**node)
                                            file_open = open(file_path + f"/commands/Cisco_{node['Node_name']}_{looptype}_profile_{create_delete}_Y1564.txt", 'w+')
                                            file_open.write(templated_command)
                                            file_open.close()
                                            print(f"**** {looptype} profile {create_delete} commands templated for {node['Node_name']}")                   
                            elif node['login']['device_type'] == 'accedian':
                                ''' prepare the accedian templates'''
                                for create_delete in ['create','delete']:
                                    for looptype in Loop_list:
                                        with open(file_path + f"/templates/Accedian_{looptype}_loop_{create_delete}_Y1564.j2",'r') as f:
                                            templated_command = Template(f.read()).render(**self.data,**node)
                                            file_open = open(file_path + f"/commands/Accedian_{node['Node_name']}_{looptype}_loop_{create_delete}_Y1564.txt", 'w+')
                                            file_open.write(templated_command)
                                            file_open.close()
                                            print(f"**** {looptype} loop {create_delete} commands templated for {node['Node_name']}")
                    ''' configure Accedian'''
                    ''' Start Y1564 traffic from Cisco'''
                    for looptype in Loop_list:
                        for create_delete in ['create','delete']:
                            for node in self.data["site_list"]:
                                if node['login']['device_type'] == 'accedian' and node['Node_name'] == acc_node:
                                    with open(file_path + '/commands/Accedian_{}_{}_loop_{}_Y1564.txt'.format(node["Node_name"],looptype,create_delete),'r') as f:
                                        output = node['connect_obj'].send_config_set(f.readlines(),cmd_verify=False)
                                        print(output)                        
                            for node in self.data["site_list"]:
                                if node['login']['device_type'] == 'cisco_xr' and node['Node_name'] == NCS_node:
                                    with open(file_path + '/commands/Cisco_{}_{}_profile_{}_Y1564.txt'.format(node["Node_name"],looptype,create_delete),'r') as f:
                                        output = node['connect_obj'].send_config_set(f.readlines())
                                        node['connect_obj'].commit()
                                        node['connect_obj'].exit_config_mode()
                                        print(output)
                                    if create_delete == 'create':
                                        time.sleep(2)
                                        command = f"show ethernet service-activation-test {node['main_interface']}.{self.data['item']}"
                                        output = node['connect_obj'].send_command(command)
                                        x = re.findall("aborted",output)
                                        if len(x) == 1:
                                            test_result[node["Node_name"]][looptype] = 'aborted'
                                            print(output)
                                        else:
                                            print("*** {} loop, Hold your breath for {} seconds, packets are on the wire".format(looptype,self.data['performance_test']*60 + 20))
                                            ''' if duration is more than 2 Mnts then disconnect the SSH connection & connect after the specified duration'''
                                            if self.data['performance_test'] > 2:
                                                self.disconnect_nodes()
                                                for i in tqdm(range(self.data['performance_test']*60 + 20),ncols=100,desc='Y1564 Progress'):
                                                    time.sleep(1)
                                                self.connect_nodes()
                                            else:
                                                for i in tqdm(range(self.data['performance_test']*60 + 20),ncols=100,desc='Y1564 Progress'):
                                                    time.sleep(1)
                                            with open(file_path + '/netconf_filters/filter_ethernet_service_test_state.j2') as f:
                                                xml_structure = Template(f.read()).render(**self.data,**node)                                                  
                                            reply_xml = node['net_conf_obj'].get(xml_structure)
                                            reply_dict = xmltodict.parse(reply_xml.xml)["rpc-reply"]["data"]["service-activation-test"]["test-statuses"]["test-status"]["phase"]["results"]
                                            output_dict = json.loads(json.dumps(reply_dict))
                                            # pprint(output_dict)                                    
                                            output = node['connect_obj'].send_command("show ethernet service-activation-test")
                                            print(output)
                                            ''' parse the Output and see FL == 0 '''
                                            x = re.findall("FL: \d+",output)
                                            try: 
                                                if int(x[0].split()[1]) == 0 and int(x[1].split()[1]) == 0 :
                                                    test_result[node["Node_name"]][looptype]= 'pass'
                                                    test_result[node["Node_name"]][f"{looptype}_CIR_TX"] = output_dict['cir']['tx-packets']
                                                    test_result[node["Node_name"]][f"{looptype}_CIR_FL"] = output_dict['cir']['frames-lost']
                                                    test_result[node["Node_name"]][f"{looptype}_EIR_TX"] = output_dict['eir']['tx-packets']
                                                    test_result[node["Node_name"]][f"{looptype}_EIR_FL"] = output_dict['eir']['frames-lost']
                                                else:
                                                    test_result[node["Node_name"]][looptype] = 'fail'
                                                    test_result[node["Node_name"]][f"{looptype}_CIR_TX"] = output_dict['cir']['tx-packets']
                                                    test_result[node["Node_name"]][f"{looptype}_CIR_FL"] = output_dict['cir']['frames-lost']
                                                    test_result[node["Node_name"]][f"{looptype}_EIR_TX"] = output_dict['eir']['tx-packets']
                                                    test_result[node["Node_name"]][f"{looptype}_EIR_FL"] = output_dict['eir']['frames-lost']
                                            except:
                                                test_result[node["Node_name"]][looptype] = 'exception'
                else:
                    test_result[node["Node_name"]] = 'not_supported'
        return test_result
    def check_QOS_configured_CIR(self):
        dict13 = {}
        for node in self.data["site_list"]:
            if 'EP' in node['side']:
                if node['login']['device_type'] == 'cisco_xr':
                    try:
                        if node['port_type'] == 'PL-type':
                            # print(f"!\n!!\n**** {node['Node_name']} # show qos interface {node['main_interface']} input")                          
                            output = node['connect_obj'].send_command("show qos interface {} input".format(node["main_interface"]))    
                        else:
                            # print(f"!\n!!\n**** {node['Node_name']} # show qos interface {node['main_interface']}.{self.data['item']} input")
                            output = node['connect_obj'].send_command("show qos interface {}.{} input".format(node["main_interface"],self.data['item']))
                        configured_cir = re.findall("Policer committed rate\s+=\s+\d+", output)
                        dict13[node['Node_name']] = configured_cir[0].split()[-1]
                    except:
                        print("*** something went Wrong, input Policy drops can not be checked")                    
                else:
                    pass
            else:
                pass
        return dict13
                

          
